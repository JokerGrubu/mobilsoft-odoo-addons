# -*- coding: utf-8 -*-
"""
MobilSoft Portal — Import/Export Controller
Excel export (cariler, ürünler, faturalar, ekstre)
"""
import logging
import io
from datetime import date

from odoo import http
from odoo.http import request
from .helpers import get_company_ids, get_default_company_id, check_record_access

_logger = logging.getLogger(__name__)


class MobilSoftImportExport(http.Controller):

    def _make_xlsx_response(self, output, filename):
        """XLSX yanıtı oluştur."""
        headers = [
            ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
            ('Content-Disposition', f'attachment; filename="{filename}"'),
        ]
        return request.make_response(output.getvalue(), headers=headers)

    # ==================== CARİLER EXPORT ====================

    @http.route('/mobilsoft/export/cariler', type='http', auth='user', website=True, sitemap=False)
    def export_cariler(self, **kwargs):
        """Cari listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/cariler?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        partners = env['res.partner'].sudo().search([
            ('company_id', 'in', company_ids + [False]),
            '|', ('customer_rank', '>', 0), ('supplier_rank', '>', 0),
        ], order='name asc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cariler'

        headers = ['Ad', 'VKN/TCKN', 'Telefon', 'E-posta', 'Adres', 'Şehir', 'Tür']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for p in partners:
            ptype = []
            if p.customer_rank > 0:
                ptype.append('Müşteri')
            if p.supplier_rank > 0:
                ptype.append('Tedarikçi')
            ws.append([
                p.name or '',
                p.vat or '',
                p.phone or '',
                p.email or '',
                p.street or '',
                p.city or '',
                '/'.join(ptype),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'cariler_{date.today()}.xlsx')

    # ==================== ÜRÜNLER EXPORT ====================

    @http.route('/mobilsoft/export/urunler', type='http', auth='user', website=True, sitemap=False)
    def export_urunler(self, **kwargs):
        """Ürün listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/urunler?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        # Products are SHARED — no company filter
        products = env['product.product'].sudo().search(
            [], order='name asc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Urunler'

        headers = ['Ad', 'Dahili Referans', 'Barkod', 'Satış Fiyatı', 'Maliyet', 'Tür', 'Kategori']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for p in products:
            ws.append([
                p.name or '',
                p.default_code or '',
                p.barcode or '',
                p.list_price,
                p.standard_price,
                'Stoğa Alınabilir' if p.type == 'product' else ('Tüketim' if p.type == 'consu' else 'Hizmet'),
                p.categ_id.name or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'urunler_{date.today()}.xlsx')

    # ==================== VERİ İÇE AKTARMA ====================

    @http.route('/mobilsoft/veri-aktarimi', type='http', auth='user', website=True, sitemap=False)
    def veri_aktarimi_page(self, **kwargs):
        """Veri içe aktarma sayfası."""
        return request.render('mobilsoft_portal.veri_aktarimi', {
            'success': kwargs.get('success', ''),
            'error': kwargs.get('error', ''),
            'imported_count': int(kwargs.get('imported_count', 0)),
            'error_lines': kwargs.get('error_lines', ''),
            'active_tab': kwargs.get('tab', 'urunler'),
        })

    @http.route('/mobilsoft/veri-aktarimi/urunler', type='http', auth='user', website=True, methods=['POST'], sitemap=False, csrf=True)
    def import_urunler(self, **post):
        """Ürün içe aktarma (Excel/XML)."""
        try:
            upload = post.get('file')
            if not upload:
                return request.redirect('/mobilsoft/veri-aktarimi?error=Dosya+seçilmedi&tab=urunler')

            filename = upload.filename.lower()
            data = upload.read()

            products_data = []
            errors = []

            if filename.endswith('.xlsx'):
                products_data, errors = self._parse_products_excel(data)
            elif filename.endswith('.xml'):
                products_data, errors = self._parse_products_xml(data)
            else:
                return request.redirect('/mobilsoft/veri-aktarimi?error=Desteklenmeyen+dosya+formatı.+xlsx+veya+xml+yükleyin&tab=urunler')

            created = 0
            env = request.env
            company_ids = get_company_ids()
            for idx, pdata in enumerate(products_data, 1):
                try:
                    vals = {
                        'name': pdata.get('name', '').strip(),
                        'barcode': pdata.get('barcode', '').strip() or False,
                        'list_price': float(pdata.get('price', 0) or 0),
                        'company_id': get_default_company_id(),
                    }
                    if not vals['name']:
                        errors.append(f"Satır {idx}: Ad boş")
                        continue
                    cat_name = pdata.get('category', '').strip()
                    if cat_name:
                        cat = env['product.category'].sudo().search([('name', '=', cat_name)], limit=1)
                        if cat:
                            vals['categ_id'] = cat.id
                    env['product.product'].sudo().create(vals)
                    created += 1
                except Exception as e:
                    errors.append(f"Satır {idx}: {e}")

            error_str = '|'.join(errors[:20]) if errors else ''
            return request.redirect(f'/mobilsoft/veri-aktarimi?success=urunler&imported_count={created}&error_lines={error_str}&tab=urunler')
        except Exception as e:
            _logger.exception("Ürün import hatası")
            return request.redirect(f'/mobilsoft/veri-aktarimi?error={e}&tab=urunler')

    @http.route('/mobilsoft/veri-aktarimi/cariler', type='http', auth='user', website=True, methods=['POST'], sitemap=False, csrf=True)
    def import_cariler(self, **post):
        """Cari içe aktarma (Excel/XML)."""
        try:
            upload = post.get('file')
            if not upload:
                return request.redirect('/mobilsoft/veri-aktarimi?error=Dosya+seçilmedi&tab=cariler')

            filename = upload.filename.lower()
            data = upload.read()

            partners_data = []
            errors = []

            if filename.endswith('.xlsx'):
                partners_data, errors = self._parse_partners_excel(data)
            elif filename.endswith('.xml'):
                partners_data, errors = self._parse_partners_xml(data)
            else:
                return request.redirect('/mobilsoft/veri-aktarimi?error=Desteklenmeyen+dosya+formatı.+xlsx+veya+xml+yükleyin&tab=cariler')

            created = 0
            env = request.env
            company_ids = get_company_ids()
            for idx, pdata in enumerate(partners_data, 1):
                try:
                    name = pdata.get('name', '').strip()
                    if not name:
                        errors.append(f"Satır {idx}: Ad boş")
                        continue
                    vals = {
                        'name': name,
                        'vat': pdata.get('vkn', '').strip() or False,
                        'phone': pdata.get('phone', '').strip() or False,
                        'email': pdata.get('email', '').strip() or False,
                        'street': pdata.get('address', '').strip() or False,
                        'customer_rank': 1,
                        'company_id': get_default_company_id(),
                    }
                    env['res.partner'].sudo().create(vals)
                    created += 1
                except Exception as e:
                    errors.append(f"Satır {idx}: {e}")

            error_str = '|'.join(errors[:20]) if errors else ''
            return request.redirect(f'/mobilsoft/veri-aktarimi?success=cariler&imported_count={created}&error_lines={error_str}&tab=cariler')
        except Exception as e:
            _logger.exception("Cari import hatası")
            return request.redirect(f'/mobilsoft/veri-aktarimi?error={e}&tab=cariler')

    @http.route('/mobilsoft/veri-aktarimi/sablon/urunler', type='http', auth='user', website=True, sitemap=False)
    def download_template_urunler(self, **kwargs):
        """Ürün import şablonu indir."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/veri-aktarimi?error=openpyxl+kurulu+değil')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ürünler'
        headers = ['Ad', 'Barkod', 'Fiyat', 'Kategori']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)
        ws.append(['Örnek Ürün', '8690000000001', '100.00', 'Genel'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, 'urun_import_sablonu.xlsx')

    @http.route('/mobilsoft/veri-aktarimi/sablon/cariler', type='http', auth='user', website=True, sitemap=False)
    def download_template_cariler(self, **kwargs):
        """Cari import şablonu indir."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/veri-aktarimi?error=openpyxl+kurulu+değil')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cariler'
        headers = ['Ad', 'VKN', 'Telefon', 'E-posta', 'Adres']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)
        ws.append(['Örnek Firma A.Ş.', '1234567890', '05551234567', 'info@firma.com', 'İstanbul'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, 'cari_import_sablonu.xlsx')

    # --- Parse helpers ---

    def _parse_products_excel(self, data):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        products = []
        errors = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            products.append({
                'name': str(row[0] or ''),
                'barcode': str(row[1] or '') if len(row) > 1 else '',
                'price': row[2] if len(row) > 2 else 0,
                'category': str(row[3] or '') if len(row) > 3 else '',
            })
        return products, errors

    def _parse_products_xml(self, data):
        import xml.etree.ElementTree as ET
        products = []
        errors = []
        try:
            root = ET.fromstring(data)
            for item in root.iter('urun'):
                products.append({
                    'name': (item.findtext('ad') or item.findtext('name') or '').strip(),
                    'barcode': (item.findtext('barkod') or item.findtext('barcode') or '').strip(),
                    'price': item.findtext('fiyat') or item.findtext('price') or 0,
                    'category': (item.findtext('kategori') or item.findtext('category') or '').strip(),
                })
        except ET.ParseError as e:
            errors.append(f"XML parse hatası: {e}")
        return products, errors

    def _parse_partners_excel(self, data):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        partners = []
        errors = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            partners.append({
                'name': str(row[0] or ''),
                'vkn': str(row[1] or '') if len(row) > 1 else '',
                'phone': str(row[2] or '') if len(row) > 2 else '',
                'email': str(row[3] or '') if len(row) > 3 else '',
                'address': str(row[4] or '') if len(row) > 4 else '',
            })
        return partners, errors

    def _parse_partners_xml(self, data):
        import xml.etree.ElementTree as ET
        partners = []
        errors = []
        try:
            root = ET.fromstring(data)
            for item in root.iter('cari'):
                partners.append({
                    'name': (item.findtext('ad') or item.findtext('name') or '').strip(),
                    'vkn': (item.findtext('vkn') or item.findtext('tckn') or '').strip(),
                    'phone': (item.findtext('telefon') or item.findtext('phone') or '').strip(),
                    'email': (item.findtext('eposta') or item.findtext('email') or '').strip(),
                    'address': (item.findtext('adres') or item.findtext('address') or '').strip(),
                })
        except ET.ParseError as e:
            errors.append(f"XML parse hatası: {e}")
        return partners, errors

    # ==================== SİPARİŞLER EXPORT ====================

    @http.route('/mobilsoft/export/siparisler', type='http', auth='user', website=True, sitemap=False)
    def export_siparisler(self, **kwargs):
        """Sipariş listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/siparisler?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        orders = env['sale.order'].sudo().search([
            ('company_id', 'in', company_ids),
        ], order='date_order desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Siparisler'

        headers = ['Sipariş No', 'Cari', 'Tarih', 'Toplam', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for o in orders:
            ws.append([
                o.name or '',
                o.partner_id.name or '',
                str(o.date_order or ''),
                o.amount_total,
                o.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'siparisler_{date.today()}.xlsx')

    # ==================== ÖDEMELER EXPORT ====================

    @http.route('/mobilsoft/export/odemeler', type='http', auth='user', website=True, sitemap=False)
    def export_odemeler(self, **kwargs):
        """Ödeme listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/odemeler?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        payments = env['account.payment'].sudo().search([
            ('company_id', 'in', company_ids),
        ], order='date desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Odemeler'

        headers = ['Ödeme No', 'Cari', 'Tarih', 'Tutar', 'Tür', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for p in payments:
            ws.append([
                p.name or '',
                p.partner_id.name or '',
                str(p.date or ''),
                p.amount,
                'Tahsilat' if p.payment_type == 'inbound' else 'Ödeme',
                p.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'odemeler_{date.today()}.xlsx')

    # ==================== İRSALİYELER EXPORT ====================

    @http.route('/mobilsoft/export/irsaliyeler', type='http', auth='user', website=True, sitemap=False)
    def export_irsaliyeler(self, **kwargs):
        """İrsaliye listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/irsaliyeler?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        pickings = env['stock.picking'].sudo().search([
            ('company_id', 'in', company_ids),
        ], order='date_done desc, scheduled_date desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Irsaliyeler'

        headers = ['İrsaliye No', 'Cari', 'Tarih', 'Kaynak', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for p in pickings:
            ws.append([
                p.name or '',
                p.partner_id.name or '',
                str(p.date_done or p.scheduled_date or ''),
                p.origin or '',
                p.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'irsaliyeler_{date.today()}.xlsx')

    # ==================== MASRAFLAR EXPORT ====================

    @http.route('/mobilsoft/export/masraflar', type='http', auth='user', website=True, sitemap=False)
    def export_masraflar(self, **kwargs):
        """Masraf listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/masraflar?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        moves = env['account.move'].sudo().search([
            ('company_id', 'in', company_ids),
            ('move_type', 'in', ['in_invoice', 'in_refund', 'entry']),
        ], order='date desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Masraflar'

        headers = ['Numara', 'Tedarikçi', 'Tarih', 'Referans', 'Tutar', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for m in moves:
            ws.append([
                m.name or '',
                m.partner_id.name or '',
                str(m.date or ''),
                m.ref or '',
                m.amount_total,
                m.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'masraflar_{date.today()}.xlsx')

    # ==================== ÇEK/SENET EXPORT ====================

    @http.route('/mobilsoft/export/cek-senet', type='http', auth='user', website=True, sitemap=False)
    def export_cek_senet(self, **kwargs):
        """Çek/Senet listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/cek-senet?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        try:
            records = env['mobilsoft.cheque.promissory'].sudo().search([
                ('company_id', 'in', company_ids),
            ], order='maturity_date desc')
        except Exception:
            return request.redirect('/mobilsoft/cek-senet?error=Çek/Senet+modülü+bulunamadı')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Cek_Senet'

        headers = ['Numara', 'Cari', 'Tutar', 'Vade Tarihi', 'Tür', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for r in records:
            ws.append([
                r.name or '',
                r.partner_id.name or '',
                r.amount,
                str(r.maturity_date or ''),
                r.type or '',
                r.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'cek_senet_{date.today()}.xlsx')

    # ==================== POS SİPARİŞLER EXPORT ====================

    @http.route('/mobilsoft/export/pos-siparisler', type='http', auth='user', website=True, sitemap=False)
    def export_pos_siparisler(self, **kwargs):
        """POS sipariş listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/pos?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        orders = env['pos.order'].sudo().search([
            ('company_id', 'in', company_ids),
        ], order='date_order desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'POS_Siparisler'

        headers = ['Sipariş No', 'Tarih', 'Müşteri', 'Toplam', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for o in orders:
            ws.append([
                o.name or '',
                str(o.date_order or ''),
                o.partner_id.name or 'Misafir',
                o.amount_total,
                o.state or '',
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'pos_siparisler_{date.today()}.xlsx')

    # ==================== FATURALAR EXPORT ====================

    @http.route('/mobilsoft/export/faturalar', type='http', auth='user', website=True, sitemap=False)
    def export_faturalar(self, tab='satis', **kwargs):
        """Fatura listesi Excel export."""
        try:
            import openpyxl
        except ImportError:
            return request.redirect('/mobilsoft/faturalar?error=openpyxl+kurulu+degil')

        env = request.env
        company_ids = get_company_ids()

        if tab == 'alis':
            move_types = ['in_invoice', 'in_refund']
        else:
            move_types = ['out_invoice', 'out_refund']

        invoices = env['account.move'].sudo().search([
            ('company_id', 'in', company_ids),
            ('move_type', 'in', move_types),
        ], order='invoice_date desc')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Faturalar'

        headers = ['Fatura No', 'Cari', 'Tarih', 'Vade', 'Ara Toplam', 'KDV', 'Toplam', 'Kalan', 'Durum']
        ws.append(headers)
        for h_cell in ws[1]:
            h_cell.font = openpyxl.styles.Font(bold=True)

        for inv in invoices:
            state_label = 'Taslak' if inv.state == 'draft' else ('İptal' if inv.state == 'cancel' else
                          ('Ödendi' if inv.payment_state == 'paid' else 'Açık'))
            ws.append([
                inv.name or 'Taslak',
                inv.partner_id.name or '',
                str(inv.invoice_date or ''),
                str(inv.invoice_date_due or ''),
                inv.amount_untaxed,
                inv.amount_tax,
                inv.amount_total,
                inv.amount_residual,
                state_label,
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return self._make_xlsx_response(output, f'faturalar_{tab}_{date.today()}.xlsx')
