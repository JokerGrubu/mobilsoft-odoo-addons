#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Luca "DETAY FİŞ LİSTESİ" Excel (xlsx) -> Odoo import-ready CSV dönüşümü.

Kullanım:
  python3 transform_journal_xlsx.py --input /joker/Mimar/sirket_verileri/Joker/2019.xlsx --year 2019

Çıktılar (default):
  /joker/Mimar/sirket_verileri/Prepared/journal_2019_prepared.csv
  /joker/Mimar/sirket_verileri/Prepared/journal_2019_odoo_import.csv
  /joker/Mimar/sirket_verileri/Prepared/journal_2019_transform_summary.txt

Not:
- Hesap kodu mapping: /joker/Mimar/sirket_verileri/JokerGrubu_Mevcut_Veriler/03_Hesap_Eslestirme.csv
- Partner/VAT mapping de aynı dosyadaki partner_vkn/partner_name alanlarından alınır.
- Eşleşmeyen hesap kodları "fallback" olarak ana hesaba (xxx000) düşürülür ve rapora yazılır.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


DEFAULT_ACCOUNT_MAP = "/joker/Mimar/sirket_verileri/JokerGrubu_Mevcut_Veriler/03_Hesap_Eslestirme.csv"
DEFAULT_OUT_DIR = "/joker/Mimar/sirket_verileri/Prepared"

SHEET_NAME = "DETAY FİŞ LİSTESİ"

META_RE = re.compile(
    r"Fiş\s*Tipi\s*/\s*Fiş\s*No\s*:\s*(?P<type>.+?)\s*/\s*(?P<no>\d+)\s*Tarih\s*:\s*(?P<date>\d{2}/\d{2}/\d{4})",
    re.IGNORECASE,
)


def norm_text(s) -> str:
    return ("" if s is None else str(s)).strip()


def load_account_mapping(path: str) -> dict:
    mapping: dict[str, dict] = {}
    with open(path, "r", encoding="utf-8", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            luca = norm_text(row.get("luca_kod") or row.get("\ufeffluca_kod"))
            if not luca:
                continue
            mapping[luca] = {
                "odoo_kod": norm_text(row.get("odoo_kod")),
                "odoo_isim": norm_text(row.get("odoo_isim")),
                "partner_vkn": norm_text(row.get("partner_vkn")),
                "partner_name": norm_text(row.get("partner_name")),
            }
    return mapping


def infer_journal_code(_fis_tipi: str) -> str:
    return "CSTL"


def to_iso_date(ddmmyyyy: str) -> str:
    try:
        return datetime.strptime(ddmmyyyy, "%d/%m/%Y").date().isoformat()
    except Exception:
        return ""


def parse_account_code(raw) -> str:
    if raw is None:
        return ""
    if isinstance(raw, (int,)):
        return str(raw)
    if isinstance(raw, float):
        # Excel often converts 501 -> 501.0
        if raw.is_integer():
            return str(int(raw))
        return str(raw)
    s = norm_text(raw)
    if s.endswith(".0") and s.replace(".", "", 1).isdigit():
        s = s[:-2]
    return s


def parse_amount(raw) -> float:
    if raw is None or raw == "":
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = norm_text(raw)
    # Turkish style: 12.420,00
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def map_account(account_map: dict, luca_code: str, luca_name: str, stats: Counter):
    rec = account_map.get(luca_code)
    if rec and rec.get("odoo_kod"):
        return rec["odoo_kod"], rec.get("odoo_isim") or luca_name, rec

    # fallback: map to main account xxx000
    m = re.match(r"^(?P<pfx>\d{3})", luca_code)
    if m:
        stats["missing_mapping_fallback_to_main"] += 1
        return f"{m.group('pfx')}000", luca_name, rec or {}

    stats["missing_mapping_skipped"] += 1
    return "", luca_name, rec or {}


def transform_xlsx(input_path: str, year: int, out_dir: str, account_map_path: str):
    account_map = load_account_mapping(account_map_path)
    stats: Counter = Counter()

    out_dir_p = Path(out_dir)
    out_dir_p.mkdir(parents=True, exist_ok=True)

    out_prepared = out_dir_p / f"journal_{year}_prepared.csv"
    out_aml = out_dir_p / f"journal_{year}_odoo_import.csv"
    out_summary = out_dir_p / f"journal_{year}_transform_summary.txt"

    prepared_header = [
        "date",
        "ref",
        "journal_code",
        "account_code",
        "account_name",
        "partner_vat",
        "partner_name",
        "label",
        "debit",
        "credit",
        "luca_code",
        "invoice_ref",
    ]
    aml_header = ["Journal/Code", "Date", "Ref", "Label", "Account/Code", "Partner", "Debit", "Credit"]

    # read_only=True bazı Luca export'larında max_column yanlış gelebiliyor; bu yüzden normal modda açıyoruz.
    wb = openpyxl.load_workbook(input_path, read_only=False, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb[wb.sheetnames[0]]
        stats["sheet_fallback"] += 1

    current_date = ""
    current_ref = ""
    current_fis_tipi = ""
    idx_code = idx_name = idx_label = idx_debit = idx_credit = None

    rows_prepared = []
    rows_aml = []
    voucher_lines = []

    def flush_voucher():
        nonlocal voucher_lines
        if not voucher_lines:
            return

        # 1) Luca hiyerarşi satırları: parent hesaplar çocuk satırların toplamı olarak tekrar eder.
        # Çocuk (alt hesap) varken parent satırı düşür (double-count engeli).
        codes = [ln.get("luca_code") or "" for ln in voucher_lines]
        parents = set()
        for c in codes:
            if not c:
                continue
            prefix = c + "."
            if any(o.startswith(prefix) for o in codes if o and o != c):
                parents.add(c)
        if parents:
            voucher_lines = [ln for ln in voucher_lines if (ln.get("luca_code") or "") not in parents]

        # Luca çıktısında aynı satırın "toplam" halleri (örn. 320 / 320.01 / 320.01.xxxxx) tekrarlar.
        # Partner hesaplarında bu tekrarlar partner alanı boş olduğundan importu da bloke eder.
        # Bu yüzden aynı (account_code, debit, credit) grubunda, partner/label dolu satırlar varken
        # tamamen boş (partner+label boş) satırları düşürürüz.
        grouped = defaultdict(list)
        for ln in voucher_lines:
            key = (ln["account_code"], ln["debit"], ln["credit"])
            grouped[key].append(ln)

        final_lines = []
        for _key, group in grouped.items():
            has_meaningful = any((g.get("partner_name") or "").strip() or (g.get("label") or "").strip() for g in group)
            if has_meaningful:
                final_lines.extend([g for g in group if (g.get("partner_name") or "").strip() or (g.get("label") or "").strip()])
            else:
                # Keep one representative (longest luca_code)
                group.sort(key=lambda g: len(g.get("luca_code") or ""), reverse=True)
                final_lines.append(group[0])

        for ln in final_lines:
            rows_prepared.append(ln)
            rows_aml.append(
                {
                    "Journal/Code": ln["journal_code"],
                    "Date": ln["date"],
                    "Ref": ln["ref"],
                    "Label": ln["label"],
                    "Account/Code": ln["account_code"],
                    "Partner": ln["partner_name"],
                    "Debit": ln["debit"],
                    "Credit": ln["credit"],
                }
            )

        voucher_lines = []

    max_col = max(6, ws.max_column or 6)
    for row in ws.iter_rows(min_col=1, max_col=max_col, values_only=True):
        cells = [norm_text(c) for c in row]
        joined = " ".join(c for c in cells if c).strip()
        if not joined:
            continue

        m = META_RE.search(joined)
        if m:
            flush_voucher()
            current_fis_tipi = m.group("type").strip()
            fis_no = m.group("no").strip()
            date_str = m.group("date").strip()
            current_date = to_iso_date(date_str)
            # ref format: YYYY/00001
            try:
                yy = datetime.strptime(date_str, "%d/%m/%Y").year
            except Exception:
                yy = year
            current_ref = f"{yy}/{fis_no}"
            idx_code = idx_name = idx_label = idx_debit = idx_credit = None
            stats["vouchers"] += 1
            continue

        up = [c.upper() for c in cells]
        if any("HESAP KODU" in c for c in up):
            for i, c in enumerate(up):
                if "HESAP KODU" in c:
                    idx_code = i
                elif "HESAP ADI" in c:
                    idx_name = i
                elif "AÇIKLAMA" in c:
                    idx_label = i
                elif c.strip() == "BORÇ" or "BORÇ" in c:
                    idx_debit = i
                elif c.strip() == "ALACAK" or "ALACAK" in c:
                    idx_credit = i
            stats["headers"] += 1
            continue

        # Stop on totals
        if "GENEL TOPLAM" in joined.upper():
            stats["totals"] += 1
            continue

        if idx_code is None or current_ref == "" or current_date == "":
            continue

        luca_code = parse_account_code(row[idx_code] if idx_code < len(row) else "")
        if not luca_code:
            continue

        luca_name = norm_text(row[idx_name]) if (idx_name is not None and idx_name < len(row)) else ""
        label = norm_text(row[idx_label]) if (idx_label is not None and idx_label < len(row)) else ""
        debit = parse_amount(row[idx_debit]) if (idx_debit is not None and idx_debit < len(row)) else 0.0
        credit = parse_amount(row[idx_credit]) if (idx_credit is not None and idx_credit < len(row)) else 0.0

        odoo_code, odoo_name, map_rec = map_account(account_map, luca_code, luca_name, stats)
        if not odoo_code:
            stats["skipped_no_account_code"] += 1
            continue

        partner_vat = norm_text(map_rec.get("partner_vkn") if map_rec else "")
        partner_name = norm_text(map_rec.get("partner_name") if map_rec else "")
        if not partner_name and (luca_code.startswith("120.01.") or luca_code.startswith("320.01.")):
            partner_name = luca_name

        journal_code = infer_journal_code(current_fis_tipi)

        voucher_lines.append(
            {
                "date": current_date,
                "ref": current_ref,
                "journal_code": journal_code,
                "account_code": odoo_code,
                "account_name": odoo_name,
                "partner_vat": partner_vat,
                "partner_name": partner_name,
                "label": label.strip(),
                "debit": f"{debit:.2f}",
                "credit": f"{credit:.2f}",
                "luca_code": luca_code,
                "invoice_ref": "",
            }
        )
        stats["lines"] += 1

    flush_voucher()

    with open(out_prepared, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=prepared_header)
        w.writeheader()
        w.writerows(rows_prepared)

    with open(out_aml, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=aml_header)
        w.writeheader()
        w.writerows(rows_aml)

    summary = []
    summary.append(f"input: {input_path}")
    summary.append(f"year: {year}")
    summary.append(f"out_prepared: {out_prepared}")
    summary.append(f"out_aml: {out_aml}")
    summary.append("")
    for k in sorted(stats.keys()):
        summary.append(f"{k}: {stats[k]}")
    out_summary.write_text("\n".join(summary) + "\n", encoding="utf-8")

    return str(out_prepared), str(out_summary)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--out-dir", default=DEFAULT_OUT_DIR)
    ap.add_argument("--account-map", default=DEFAULT_ACCOUNT_MAP)
    args = ap.parse_args()

    transform_xlsx(args.input, args.year, args.out_dir, args.account_map)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
