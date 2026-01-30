# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import csv
import io
from datetime import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom


class LucaExportMixin(models.AbstractModel):
    """
    Mixin class providing Luca export utilities
    """
    _name = 'luca.export.mixin'
    _description = 'Luca Export Utilities'

    def _get_luca_account_code(self, account):
        """
        Odoo hesap kodunu Luca formatına çevir
        Luca genelde 3 seviyeli hesap planı kullanır: XXX.XX.XXX
        """
        if not account:
            return ''
        # Odoo'da hesap kodu zaten uygun formatta olabilir
        return account.code or ''

    def _format_luca_date(self, date_value):
        """
        Tarihi Luca formatına çevir (DD.MM.YYYY)
        """
        if not date_value:
            return ''
        if isinstance(date_value, str):
            date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
        return date_value.strftime('%d.%m.%Y')

    def _format_luca_amount(self, amount):
        """
        Tutarı Luca formatına çevir (nokta ondalık ayırıcı)
        """
        if not amount:
            return '0.00'
        # Luca nokta kullanır, virgül değil
        return '{:.2f}'.format(amount)

    def _sanitize_text(self, text):
        """
        Metni CSV/XML için temizle
        """
        if not text:
            return ''
        # Satır sonlarını ve özel karakterleri temizle
        text = str(text).replace('\n', ' ').replace('\r', ' ')
        text = text.replace('"', "'").replace(';', ',')
        return text.strip()


class AccountAccount(models.Model):
    """
    Hesap Planı için Luca export özellikleri
    """
    _inherit = 'account.account'

    def export_to_luca_xml(self):
        """
        Hesap planını Luca XML formatında export et
        Logo uyumlu XML formatı kullanılır
        """
        root = ET.Element('HESAPPLANI')
        root.set('VERSION', '1.0')

        for account in self:
            hesap = ET.SubElement(root, 'HESAP')
            ET.SubElement(hesap, 'HESAPKODU').text = account.code or ''
            ET.SubElement(hesap, 'HESAPADI').text = account.name or ''
            ET.SubElement(hesap, 'HESAPTURU').text = self._get_account_type_for_luca(account)
            ET.SubElement(hesap, 'USTHESAP').text = self._get_parent_code(account)
            ET.SubElement(hesap, 'DOVIZCINSI').text = account.currency_id.name if account.currency_id else 'TRY'

        # Pretty print XML
        xml_str = ET.tostring(root, encoding='unicode')
        dom = minidom.parseString(xml_str)
        return dom.toprettyxml(indent='  ', encoding='UTF-8')

    def _get_account_type_for_luca(self, account):
        """
        Odoo hesap tipini Luca hesap tipine çevir
        """
        type_mapping = {
            'asset_receivable': 'A',  # Alacak
            'asset_cash': 'A',        # Kasa
            'asset_current': 'A',     # Dönen Varlık
            'asset_non_current': 'A', # Duran Varlık
            'asset_fixed': 'A',       # Sabit Varlık
            'liability_payable': 'B', # Borç
            'liability_current': 'B', # Kısa Vadeli Borç
            'liability_non_current': 'B', # Uzun Vadeli Borç
            'equity': 'O',            # Özkaynak
            'income': 'G',            # Gelir
            'income_other': 'G',      # Diğer Gelir
            'expense': 'G',           # Gider
            'expense_direct_cost': 'G', # Direkt Maliyet
            'expense_depreciation': 'G', # Amortisman
            'off_balance': 'N',       # Nazım
        }
        return type_mapping.get(account.account_type, 'A')

    def _get_parent_code(self, account):
        """
        Üst hesap kodunu bul
        Örnek: 120.01.001 -> 120.01
        """
        code = account.code or ''
        if '.' in code:
            parts = code.rsplit('.', 1)
            return parts[0]
        elif len(code) > 3:
            return code[:-3]
        return ''


class AccountMove(models.Model):
    """
    Muhasebe Fişleri için Luca export özellikleri
    """
    _inherit = 'account.move'

    def export_to_luca_csv(self, include_header=True):
        """
        Muhasebe fişlerini Luca CSV formatında export et

        Luca Zorunlu Alanlar:
        - Fiş No
        - Fiş Tarihi
        - Hesap Kodu
        - Borç
        - Alacak

        Opsiyonel:
        - Açıklama
        - Belge No
        - Belge Tarihi
        """
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        if include_header:
            writer.writerow([
                'Fiş No',
                'Fiş Tarihi',
                'Hesap Kodu',
                'Borç',
                'Alacak',
                'Açıklama',
                'Belge No',
                'Belge Tarihi',
            ])

        mixin = self.env['luca.export.mixin']

        for move in self:
            fis_no = move.name or ''
            fis_tarihi = mixin._format_luca_date(move.date)
            belge_no = move.ref or ''
            belge_tarihi = mixin._format_luca_date(move.invoice_date) if hasattr(move, 'invoice_date') and move.invoice_date else fis_tarihi

            for line in move.line_ids:
                if not line.account_id:
                    continue

                hesap_kodu = line.account_id.code or ''
                borc = mixin._format_luca_amount(line.debit)
                alacak = mixin._format_luca_amount(line.credit)
                aciklama = mixin._sanitize_text(line.name or move.narration or '')

                writer.writerow([
                    fis_no,
                    fis_tarihi,
                    hesap_kodu,
                    borc,
                    alacak,
                    aciklama,
                    belge_no,
                    belge_tarihi,
                ])

        return output.getvalue()

    def export_to_luca_excel(self):
        """
        Muhasebe fişlerini Luca Excel formatında export et
        openpyxl kullanarak .xlsx dosyası oluşturur
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            raise UserError(_('Excel export için openpyxl kütüphanesi gerekli. pip install openpyxl'))

        wb = Workbook()
        ws = wb.active
        ws.title = 'Muhasebe Fişleri'

        # Header
        headers = ['Fiş No', 'Fiş Tarihi', 'Hesap Kodu', 'Borç', 'Alacak', 'Açıklama', 'Belge No', 'Belge Tarihi']
        header_font = Font(bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font

        mixin = self.env['luca.export.mixin']
        row = 2

        for move in self:
            fis_no = move.name or ''
            fis_tarihi = mixin._format_luca_date(move.date)
            belge_no = move.ref or ''
            belge_tarihi = mixin._format_luca_date(move.invoice_date) if hasattr(move, 'invoice_date') and move.invoice_date else fis_tarihi

            for line in move.line_ids:
                if not line.account_id:
                    continue

                ws.cell(row=row, column=1, value=fis_no)
                ws.cell(row=row, column=2, value=fis_tarihi)
                ws.cell(row=row, column=3, value=line.account_id.code or '')
                ws.cell(row=row, column=4, value=line.debit or 0)
                ws.cell(row=row, column=5, value=line.credit or 0)
                ws.cell(row=row, column=6, value=mixin._sanitize_text(line.name or move.narration or ''))
                ws.cell(row=row, column=7, value=belge_no)
                ws.cell(row=row, column=8, value=belge_tarihi)
                row += 1

        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 12

        # BytesIO'ya kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ResPartner(models.Model):
    """
    Cari bilgileri için Luca export özellikleri
    """
    _inherit = 'res.partner'

    def export_to_luca_csv(self, include_header=True):
        """
        Cari bilgilerini Luca CSV formatında export et
        VKN/TCKN, Adres, İletişim bilgileri
        """
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

        if include_header:
            writer.writerow([
                'Cari Kodu',
                'Cari Adı',
                'VKN/TCKN',
                'Vergi Dairesi',
                'Adres',
                'İl',
                'İlçe',
                'Telefon',
                'E-posta',
                'Tip',
            ])

        mixin = self.env['luca.export.mixin']

        for partner in self:
            # VKN/TCKN normalize et
            vkn = ''
            if partner.vat:
                vkn = ''.join(filter(str.isdigit, partner.vat))

            # Tip belirle
            if partner.is_company:
                tip = 'Şirket'
            elif len(vkn) == 11:
                tip = 'Şahıs'
            else:
                tip = 'Diğer'

            writer.writerow([
                partner.ref or partner.id,
                mixin._sanitize_text(partner.name),
                vkn,
                mixin._sanitize_text(partner.company_registry) if hasattr(partner, 'company_registry') else '',
                mixin._sanitize_text(partner.street or ''),
                mixin._sanitize_text(partner.state_id.name if partner.state_id else ''),
                mixin._sanitize_text(partner.city or ''),
                partner.phone or partner.mobile or '',
                partner.email or '',
                tip,
            ])

        return output.getvalue()
